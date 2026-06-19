'''
comMS motif extension: resolve foreground/background protein sets and export as FASTA
'''

# -- Import external dependencies
import sys, tomllib
from dataclasses import dataclass
from openpyxl import load_workbook
from pathlib import Path

# -- Import internal comMS logMsg class
from comms.utils.log import logMsg

# Import function to read FASTA files
from comms.utils.fasta import readFasta, writeFasta

# -- Define variables for inputs
SIGNALP_TSV = "signalp_predictions.tsv"               # SignalP output

# -- Define dataclass ResolvedInput for holding the resolved input fields
@dataclass
class ResolvedInput:
    foreground_ids: list[str]
    background_ids: list[str]
    fasta_path: Path
    label: str                 # contrast name, or 'standalone'
    background_source: str     # human-readable description for context

# -- read_id_list: returns a list of strings corresponding to protein ids
def read_id_list(path: Path) -> list[str]:
    out: list[str] = []
    with path.open() as fh:
        for line in fh:
            token = line.strip().split('\t')[0]
            if token and not token.startswith('#'):
                out.append(token)
    return out

# -- _read_xlsx_rows: returns a list of dictionaries corresponding to rows in Excel spreadsheer
def _read_xlsx_rows(path: Path, sheet: str | None) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else '' for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:]]

# -- resolve_search_fasta: returns a Path to FASTA database used
def resolve_search_fasta(explicit: Path | None, experiment_dir: Path | None) -> Path:
    if explicit is not None:
        if not explicit.exists():
            raise FileNotFoundError(f'--fasta not found: {explicit}')
        return explicit
    if experiment_dir is None:
        raise ValueError('No --fasta given and no source directory to auto-resolve from')
    ctx = experiment_dir / 'experiment.toml'   # experiment context
    if not ctx.exists():
        raise FileNotFoundError(f'Could not auto-resolve the search FASTA, expected experiment context at {ctx}; pass --fasta explicitly')
    with ctx.open('rb') as fh:
        table = tomllib.load(fh)
    fasta = table.get('database', {}).get('fasta')
    if not fasta:
        raise KeyError(f'No FASTA path recorded in {ctx}, pass --fasta explicitly')
    return Path(fasta)

# -- resolve_from_report_da: returns tuple of list of strings/strings corresponding to resolved foreground and background protein sets
def resolve_from_report_da(
    cfg,
    experiment_dir: Path,
    contrast: str,
    background_choice: str,
    background_ids_file: Path | None,
) -> tuple[list[str], list[str], str]:
    # Import input schema from config
    workbook = Path(cfg.get('input', {}).get('da_workbook'))
    sheet_per_contrast = cfg.get('input', {}).get('da_sheet_per_contrast')
    protein_id = cfg.get('input', {}).get('da_col_protein_id')
    direction_col = cfg.get('input', {}).get('da_col_direction')
    fg_directions = cfg.get('input', {}).get('da_foreground_directions')
    bg_directions = cfg.get('input', {}).get('da_background_directions')
    # Read workbook
    rows = _read_xlsx_rows(experiment_dir / 'comms/results/report' / workbook, contrast) if sheet_per_contrast else [r for r in rows if r.get('contrast') == contrast]
    # Get foreground proteins
    foreground = [r[protein_id] for r in rows if r.get(direction_col) in fg_directions]
    # Get background proteins
    if background_choice == 'custom':
        if background_ids_file is None:
            raise ValueError('--background custom requires --background-ids')
        background = read_id_list(background_ids_file)
        source = f'custom ({background_ids_file.name})'
    elif background_choice == 'organism-proteome':
        background = []     # signalled empty; FASTA writer uses all non-foreground IDs
        source = 'organism reference proteome'
    else:  # detected-unchanged (default)
        background = [r[protein_id] for r in rows if r.get(direction_col) == bg_directions]
        source = 'detected-unchanged in same fraction'
    return foreground, background, source

# -- resolve_from_secondary: returns 
def resolve_from_secondary(
    cfg,
    experiment_dir: Path,
    background_choice: str,
    background_ids_file: Path | None,
) -> tuple[list[str], list[str], str]:
    # Import input schema from config
    workbook = Path(cfg.get('input', {}).get('secondary_workbook'))
    sheet = cfg.get('input', {}).get('secondary_sheet')
    protein_id = cfg.get('input', {}).get('secondary_col_protein_id')
    fraction = cfg.get('input', {}).get('secondary_fraction')
    fg_fractions = cfg.get('input', {}).get('secondary_foreground_fractions')
    # Read workbook
    rows = _read_xlsx_rows(experiment_dir / 'comms/results/report' / workbook, sheet)
    # Get foreground proteins
    foreground = [r[protein_id] for r in rows if r.get(fraction) in fg_fractions]
    # Get background proteins
    if background_choice == 'custom':
        if background_ids_file is None:
            raise ValueError('--background custom requires --background-ids')
        background = read_id_list(background_ids_file)
        source = f'custom ({background_ids_file.name})'
    else:
        background = []
        source = 'Secondary species reference proteome'
    return foreground, background, source

# -- read_sp_cleavage: returns dictionary of strings: ints which map protein_id -> signal peptide cleavage positiosn
def read_sp_cleavage(cfg, annotate_dir: Path | None) -> dict[str, int]:
    if annotate_dir is None:
        return {}
    tsv = annotate_dir / SIGNALP_TSV
    if not tsv.exists():
        logMsg.warn(f'SignalP predictions {tsv} not found')
        return {}
    table: dict[str, int] = {}
    with tsv.open() as fh:
        header = fh.readline().rstrip("\n").split("\t")
        try:
            id_i = header.index(cfg.get('input', {}).get('signalp_protein_id'))
            cs_i = header.index(cfg.get('input', {}).get('signalp_cleavage_site'))
        except ValueError:
            logMsg.warn(f'Unexpected columns in {tsv}')
            return {}
        for line in fh:
            parts = line.rstrip('\n').split('\t')
            try:
                table[parts[id_i]] = int(parts[cs_i])
            except (IndexError, ValueError):
                continue
    return table

# -- apply_window: returns string corresponding to protein sequence window
def apply_window(seq: str, window: str, sp_cleavage: int | None) -> str:
    if window == 'full':
        return seq
    if window == 'n_terminal_60':
        return seq[:60]
    if window == 'n_terminal_after_sp':
        if sp_cleavage is None:
            logMsg.warn('No SP cleavage site, using full window')
            return seq
        return seq[sp_cleavage: sp_cleavage + 60]
    raise ValueError(f'Unknown window: {window!r}')

# -- resolve_inputs: returns ResolvedInput
def resolve_inputs(
    *,
    config,
    experiment_dir: Path,
    mode: str,                 # 'standalone' | 'report-da' | 'report-secondary' | 'quantify'
    foreground_ids: Path | None,
    background_ids: Path | None,
    contrast: str | None,
    background_choice: str,
    fasta: Path | None,
) -> ResolvedInput:
    if mode == 'standalone':
        if foreground_ids is None:
            raise ValueError('Standalone mode requires --foreground-ids')
        fg = read_id_list(foreground_ids)
        bg = read_id_list(background_ids) if background_ids else []
        fasta_path = resolve_search_fasta(fasta, None)
        return ResolvedInput(fg, bg, fasta_path, "standalone", 'custom' if background_ids else 'none')
    if mode == 'report-da':
        if contrast is None:
            raise ValueError('report-da mode requires --contrast')
        fg, bg, source = resolve_from_report_da(config, experiment_dir, contrast, background_choice, background_ids)
        fasta_path = resolve_search_fasta(fasta, experiment_dir)
        return ResolvedInput(fg, bg, fasta_path, contrast, source)
    if mode == 'report-secondary':
        fg, bg, source = resolve_from_secondary(config, experiment_dir, background_choice, background_ids)
        fasta_path = resolve_search_fasta(fasta, experiment_dir)
        return ResolvedInput(fg, bg, fasta_path, "secondary_species_candidates", source)
    raise NotImplementedError(f'Input mode {mode!r} not implemented')